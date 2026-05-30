"""Planilhas XLSX (openpyxl) — cabeçalho com cor, bordas, título opcional, painéis fixos."""
from __future__ import annotations

import io
from typing import Any, List, Sequence, Union

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from vereda_backend.core.text_polish import strip_llm_markdown_artifacts


def build_xlsx_bytes(
    sheet_title: str,
    rows: Sequence[Sequence[Union[str, int, float, None]]],
    header: bool = True,
    *,
    document_title: str | None = None,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_title or "Dados")[:31]

    thin = Side(style="thin", color="CBD5E1")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    if not rows:
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    extra_top = 0
    max_cols = max(len(list(r)) for r in rows) if rows else 1
    doc_title = strip_llm_markdown_artifacts(document_title or "").strip()

    if doc_title:
        extra_top = 1
        mc = max(1, max_cols)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=mc)
        c = ws.cell(row=1, column=1, value=doc_title)
        c.font = Font(bold=True, size=14, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2563EB")
        c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        for col in range(1, max_cols + 1):
            ws.cell(row=1, column=col).border = border_all
        ws.row_dimensions[1].height = 36

    bold = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="1E293B")
    align = Alignment(vertical="center", wrap_text=True)

    r0 = list(rows[0]) if rows else []
    header_row_idx = 1 + extra_top

    if header and r0:
        for col, val in enumerate(r0, start=1):
            cell = ws.cell(row=header_row_idx, column=col, value=val)
            cell.font = bold
            cell.fill = fill
            cell.alignment = align
            cell.border = border_all
        start = 1

    body_rows: List[Sequence[Any]] = list(rows[1:]) if header else list(rows)
    offset = header_row_idx + 1 if header else header_row_idx
    
    num_align = Alignment(vertical="top", horizontal="right", wrap_text=True)
    txt_align = Alignment(vertical="top", wrap_text=True)
    for i, row in enumerate(body_rows):
        for j, val in enumerate(row):
            c = ws.cell(row=i + offset, column=j + 1, value=val)
            is_num = isinstance(val, (int, float)) and not isinstance(val, bool)
            c.alignment = num_align if is_num else txt_align
            c.border = border_all
            if i % 2 == 1:
                c.fill = PatternFill("solid", fgColor="F1F5F9")

    for col in range(1, max_cols + 1):
        letter = get_column_letter(col)
        max_len = 10
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=col).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        # Limitar largura para não estourar visualmente no Excel
        ws.column_dimensions[letter].width = min(55, max(12, max_len + 2))

    ws.freeze_panes = ws.cell(row=header_row_idx + 1, column=1)

    if header and r0 and ws.max_row >= header_row_idx:
        end_col = get_column_letter(max_cols)
        ws.auto_filter.ref = (
            f"A{header_row_idx}:{end_col}{ws.max_row}"
        )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
